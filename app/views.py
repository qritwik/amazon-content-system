from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from lxml import html
import csv,os,json
import requests
import pandas as pd
from . import forms
from .models import *
from bs4 import BeautifulSoup
import http.client
import json
import ast
from django.db.models import F
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import colors




def manager_user(request,email):
    email=email

    data1 = asinDetail.objects.filter(email=email)
    total_asin = data1.count()

    data2 = asinDetail.objects.filter(email=email).filter(status=True)
    done_asin = data2.count()

    data3 = empDetail.objects.get(email=email)

    context = {
        'email':email,
        'total_asin':total_asin,
        'done_asin':done_asin,
        'data3':data3
    }

    if request.method == 'POST':
        c=0
        user_asin_c = request.POST.get('user_asin')
        user_asin_c_n = int(user_asin_c)

        data1 = asinDetail.objects.filter(status=False).filter(extracted=True).filter(email__isnull=True)
        for i1 in data1:
            if(c<user_asin_c_n):
                i1.email = email
                i1.save()
                c=c+1
            else:
                c=0
                return HttpResponseRedirect('/user/'+email)




    return render(request,'manager_user.html',context=context)





def message(request):
    return render(request,'message.html')


def fetch_asin(request):
    data1 = asinDetail.objects.all()
    total_asin = data1.count()

    data2 = asinDetail.objects.filter(extracted=True)
    fetch_asin = data2.count()

    if fetch_asin == total_asin:
        return HttpResponseRedirect('/manager')


    form1 = forms.form_oldDetailAmazon()

    if request.method == 'POST':
        if asinDetail.objects.filter(extracted=False).exists():
            data3 = asinDetail.objects.filter(extracted=False)
            for i in data3:
                form1 = forms.form_oldDetailAmazon(request.POST)
                sendme = asinDetail.objects.get(asin=i)
                i = i.asin
                url = "http://www.amazon.in/dp/"+i
                data4 = parse(url)
                try:
                    data4['asin']=i

                except TypeError:
                    continue

                if form1.is_valid():
                    print(i)
                    obj = form1.save(commit=False)

                    obj.old_name = data4['NAME']
                    obj.old_url = data4['URL']
                    obj.old_desc = data4['DESC']
                    obj.old_brand = data4['BRAND']
                    obj.old_product_desc = data4['PRODUCT_DESC']
                    obj.old_from_manufacture_h = data4['H1']
                    obj.old_from_manufacture_p = data4['P1']

                    obj.asin = i

                    obj.save()

                    if sendme.extracted == False:
                        sendme.extracted=True
                        sendme.save()
                    return HttpResponseRedirect('/fetch_asin')
                else:
                    print(form1.errors)











    context = {
        'total_asin':total_asin,
        'fetch_asin':fetch_asin
    }

    return render(request,'fetch_asin.html',context=context)



def index(request):
    form5 = forms.form_empDetail()
    if(request.method=='POST'):
        form5 = forms.form_empDetail(request.POST)
        email = request.POST.get('email1')
        password = request.POST.get('pass1')

        conn = http.client.HTTPConnection("textmercato.com:4523")
        payload = "email="+email+"&password="+password
        headers = {
                'authorization': "Basic YWRtaW4tZGV2OmFkbWluUGFzcy0tMTIz",
                'content-type': "application/x-www-form-urlencoded",
                'cache-control': "no-cache",
                'postman-token': "5ac9f9ac-261e-c0ff-ee9b-65064632ba6e"
                }

        conn.request("POST", "/signin", payload, headers)

        res = conn.getresponse()
        data = res.read()

        data1 = data.decode("utf-8")
        data2 = ast.literal_eval(data1)
        status = data2["status"]
        first_name = data2["data"]["first_name"]
        last_name = data2["data"]["last_name"]
        name = first_name+" "+last_name
        request.session['name'] = name
        request.session['email'] = email

        roles = data2["data"]["roles"][0]
        email = data2["data"]["email"]

        try:
            if empDetail.objects.get(email=email):
                print("User already exists")

            else:
                if form5.is_valid():
                    obj5 = form5.save(commit=False)
                    obj5.name=name
                    obj5.email=email
                    obj5.roles=roles
                    obj5.save()
        except empDetail.DoesNotExist:
            #user is visiting first time
            if form5.is_valid():
                obj5 = form5.save(commit=False)
                obj5.name=name
                obj5.email=email
                obj5.roles=roles
                obj5.save()












        if status == "success":
            if roles == "admin":
                return HttpResponseRedirect('/manager')
            else:
                return HttpResponseRedirect('/detail')



        else:
            return HttpResponseRedirect('/')


    return render(request,'index.html')


def manager(request):
    c1 = 0
    c2 = 2
    c3 = 3
    data1 = asinDetail.objects.all()
    total_asin = data1.count()

    data2 = asinDetail.objects.filter(status=True)
    done_asin = data2.count()

    data3 = empDetail.objects.filter(~Q(roles="admin"))
    total_user = data3.count()

    data4 = asinDetail.objects.filter(extracted=True)
    asin_fetch = data4.count()

    data5 = asinDetail.objects.filter(~Q(email__isnull=True))
    asin_allocated = data5.count()

    if(request.method == 'POST'):


        book = Workbook()
        sheet1 = book['Sheet']

        book.create_sheet()
        sheet2 = book['Sheet1']

        book.create_sheet()
        sheet3 = book['Sheet2']

        sheet1.title = "Title"
        sheet1.column_dimensions['A'].width = 18
        sheet1.column_dimensions['B'].width = 70
        sheet1.column_dimensions['C'].width = 70


        sheet1['A1'] = "ASIN"
        sheet1['B1'] = "Current Title"
        sheet1['C1'] = "Revised Title"

        data7 = oldProductDetail.objects.all()
        for row1 in data7:
            sheet1['A'+str(c2)] = row1.asin
            sheet1['B'+str(c2)] = row1.current_Title
            sheet1['C'+str(c2)] = row1.revised_Title
            c2 = c2 + 1




        sheet2.title = "Bullet points"
        sheet2.column_dimensions['A'].width = 18
        sheet2.column_dimensions['B'].width = 70
        sheet2.column_dimensions['C'].width = 70
        sheet2.column_dimensions['D'].width = 70
        sheet2.column_dimensions['E'].width = 70
        sheet2.column_dimensions['F'].width = 70
        sheet2.column_dimensions['G'].width = 70
        sheet2.column_dimensions['H'].width = 70
        sheet2.column_dimensions['I'].width = 70
        sheet2.column_dimensions['J'].width = 70
        sheet2.column_dimensions['K'].width = 70
        





        sheet2['A1'] = "ASIN"
        sheet2['B1'] = "Title"

        sheet2['C1'] = "Bullet Point 1 (Resolution and Refresh)"
        sheet2['D1'] = "Bullet Point 2 (Display)"
        sheet2['E1'] = "Bullet Point 3 (Smart TV Features) Optional"
        sheet2['F1'] = "Bullet Point 4 (Ports Connectivity)"
        sheet2['G1'] = "Bullet Point 5 (Sound)"
        sheet2['H1'] = "Bullet Point 6 (Installation)"
        sheet2['I1'] = "Bullet Point 7 (Warranty)"
        sheet2['J1'] = "Bullet Point 8 (Additional Information) Optional"
        sheet2['K1'] = "Comments"
        data8 = newProductDetail.objects.all()
        for row2 in data8:
            sheet2['A'+str(c3)] = row2.asin
            sheet2['B'+str(c3)] = row2.title
            sheet2['C'+str(c3)] = row2.bp1
            sheet2['D'+str(c3)] = row2.bp2
            sheet2['E'+str(c3)] = row2.bp3
            sheet2['F'+str(c3)] = row2.bp4
            sheet2['G'+str(c3)] = row2.bp5
            sheet2['H'+str(c3)] = row2.bp6
            sheet2['I'+str(c3)] = row2.bp7
            sheet2['J'+str(c3)] = row2.bp8
            sheet2['K'+str(c3)] = row2.comments

            c3=c3+1










        sheet3.title = "Feature Image"
        sheet3.column_dimensions['A'].width = 30
        sheet3.column_dimensions['B'].width = 48
        sheet3['B1'] = "TV Features"
        ft = Font(color=colors.WHITE,bold=True)
        sheet3['B1'].font=ft
        sheet3['A1'].fill = PatternFill(bgColor="00000000", fill_type = "solid")
        sheet3['B1'].fill = PatternFill(bgColor="00000000", fill_type = "solid")
        data6 = featureImage.objects.all()
        for row in data6:


            sheet3['A'+str(eval("2+c1*9"))] = row.asin
            sheet3['A'+str(eval("3+c1*9"))] = "Screen size & Resolution"
            sheet3['B'+str(eval("3+c1*9"))] = row.screen_size_resolution
            sheet3['A'+str(eval("4+c1*9"))] = "Connectivity ports"
            sheet3['B'+str(eval("4+c1*9"))] = row.connectivity_ports
            sheet3['A'+str(eval("5+c1*9"))] = "Display and Refresh rate"
            sheet3['B'+str(eval("5+c1*9"))] = row.display_and_refresh_rate
            sheet3['A'+str(eval("6+c1*9"))] = "Sound output"
            sheet3['B'+str(eval("6+c1*9"))] = row.sound_output
            sheet3['A'+str(eval("7+c1*9"))] = "Smart TV"
            sheet3['B'+str(eval("7+c1*9"))] = row.smart_tv
            sheet3['A'+str(eval("8+c1*9"))] = "Warranty*"
            sheet3['B'+str(eval("8+c1*9"))] = row.warranty
            sheet3['A'+str(eval("9+c1*9"))] = "*Please check warranty information box below for details on warranty"

            c1=c1+1

        book.save("Amazon.xlsx")














    context = {
        'total_asin':total_asin,
        'asin_allocated':asin_allocated,
        'done_asin':done_asin,
        'total_user':total_user,
        'asin_fetch':asin_fetch,
        'data3':data3
        }
    return render(request,'manager.html',context=context)



def parse(url):
    h=[]
    p=[]


    headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.90 Safari/537.36'}
    r = requests.get(url,headers=headers)
    for i in range(20):
        try:

            soup = BeautifulSoup(r.content, 'html5lib')
            desc_list=[]

            doc = html.fromstring(r.content)
            XPATH_NAME = '//h1[@id="title"]//text()'
            # XPATH_SALE_PRICE = '//span[contains(@id,"ourprice") or contains(@id,"saleprice")]/text()'
            # XPATH_ORIGINAL_PRICE = '//td[contains(text(),"List Price") or contains(text(),"M.R.P") or contains(text(),"Price")]/following-sibling::td/text()'
            # XPATH_CATEGORY = '//a[@class="a-link-normal a-color-tertiary"]//text()'
            # XPATH_AVAILABILITY = '//div[@id="availability"]//text()'


            RAW_NAME = doc.xpath(XPATH_NAME)
            # RAW_SALE_PRICE = doc.xpath(XPATH_SALE_PRICE)
            # RAW_CATEGORY = doc.xpath(XPATH_CATEGORY)
            # RAW_ORIGINAL_PRICE = doc.xpath(XPATH_ORIGINAL_PRICE)
            # RAw_AVAILABILITY = doc.xpath(XPATH_AVAILABILITY)


            NAME = ' '.join(''.join(RAW_NAME).split()) if RAW_NAME else None
            # SALE_PRICE = ' '.join(''.join(RAW_SALE_PRICE).split()).strip() if RAW_SALE_PRICE else None
            # CATEGORY = ' > '.join([i.strip() for i in RAW_CATEGORY]) if RAW_CATEGORY else None
            # ORIGINAL_PRICE = ''.join(RAW_ORIGINAL_PRICE).strip() if RAW_ORIGINAL_PRICE else None
            # AVAILABILITY = ''.join(RAw_AVAILABILITY).strip() if RAw_AVAILABILITY else None

            try:
                if soup.find('div',{'class':'a-section a-spacing-medium a-spacing-top-small'}):
                    data1 = soup.find('div',{'class':'a-section a-spacing-medium a-spacing-top-small'})
                    if data1.find_all('span',{'class':'a-list-item'}):
                        for data2 in data1.find_all('span',{'class':'a-list-item'}):
                            desc_list.append(data2.text)

                if soup.find('td',{'class':'value'}):
                    data2 = soup.find('td',{'class':'value'})
                    brand = data2.text
                else:
                    brand="No Info Available"

            except TypeError or ValueError:
                continue

            if soup.find('div',{'id':'productDescription'}):
                data3 = soup.find('div',{'id':'productDescription'})
                data4 = data3.find('p')
                productDescription = data4.text
            else:
                productDescription="No Info Available"



            if soup.find('div',{'id':'aplus_feature_div'}):
                data6 = soup.find('div',{'id':'aplus_feature_div'})
                if data6.find('div',{'class':'a-section a-spacing-extra-large bucket'}):
                    data7 = data6.find('div',{'class':'a-section a-spacing-extra-large bucket'})
                    if data7.find('div',{'class':'aplus-v2 desktop celwidget'}):
                        data8 = data7.find('div',{'class':'aplus-v2 desktop celwidget'})
                        for data10 in data8.find_all('h4',{'class':'a-spacing-mini'}):
                            h.append(data10.text)
                        for data11 in data8.find_all('p',{'class':'a-size-small'}):
                            p.append(data11.text)



            # if soup.find('div',{'id':'dpx-aplus-product-description_feature_div'}):
            #     data5 = soup.find('div',{'id':'dpx-aplus-product-description_feature_div'})
            #     if data5.find('div',{'id':'aplus_feature_div'}):
            #         data6 = data5.find('div',{'id':'aplus_feature_div'})
            #         if data6.find('div',{'class':'a-section a-spacing-extra-large bucket'}):
            #             data7 = data6.find('div',{'class':'a-section a-spacing-extra-large bucket'})
            #             if data7.find('div',{'class':'aplus-v2 desktop celwidget'}):
            #                 data8 = data7.find('div',{'class':'aplus-v2 desktop celwidget'})
            #                 for data10 in data8.find_all('h4',{'class':'a-spacing-mini'}):
            #                     h.append(data10.text)
            #                 for data11 in data8.find_all('p',{'class':'a-size-small'}):
            #                     p.append(data11.text)























            # if not ORIGINAL_PRICE:
            #     ORIGINAL_PRICE = SALE_PRICE
            if not NAME :
                NAME=''

            print(h)

            #over = zip(h,p)

            data = {
                    'NAME':NAME,
                    # 'SALE_PRICE':SALE_PRICE,
                    # 'CATEGORY':CATEGORY,
                    # 'ORIGINAL_PRICE':ORIGINAL_PRICE,
                    # 'AVAILABILITY':AVAILABILITY,
                    'URL':url,
                    'DESC':desc_list,
                    'BRAND':brand,
                    'PRODUCT_DESC':productDescription,
                    'H1':h,
                    'P1':p
                    }

            return data
        except Exception as e:
            print(e)



#
# def detail(request):
#     name = request.session.get('name')
#     email = request.session.get('email')
#
#     if asinDetail.objects.filter(status=False).filter(email=email).filter(extracted=True).exists():
#         data2 = asinDetail.objects.filter(status=False, extracted=True, email=email)
#         for i in data2:
#             i = i.asin
#             print(i)
#             data3 = oldDetailAmazon.objects.get(asin=i)
#             form1 = forms.form_newProductDetail()
#             form2 = forms.form_oldProductDetail()
#             context = {
#                 'data3':data3,
#                 'form1':form1,
#                 'form2':form2,
#                 'name':name
#                 }
#
#             if request.method == 'POST':
#                 newtitle = request.POST.get('newtitle')
#                 form1 = forms.form_newProductDetail(request.POST)
#                 form2 = forms.form_oldProductDetail(request.POST)
#                 sendme = asinDetail.objects.get(asin=i)
#
#                 if form1.is_valid() and form2.is_valid():
#                     obj = form1.save(commit=False)
#                     obj.asin=i
#                     obj.save()
#
#                     obj1 = form2.save(commit=False)
#                     obj1.asin=i
#                     obj1.current_Title=data3.old_name
#                     obj1.revised_Title=form1.cleaned_data['title']
#                     obj1.save()
#
#                     data7 = empDetail.objects.get(email=email)
#                     data7.asin_done_c = F('asin_done_c') + 1
#                     data7.save()
#
#                     if sendme.status == False:
#                         sendme.status=True
#                         sendme.save()
#
#
#
#
#
#
#
#
#                 # with open('asinn.csv', 'r') as fin:
#                 #     data = fin.read().splitlines(True)
#                 # with open('asinn.csv', 'w') as fout:
#                 #     fout.writelines(data[1:])
#                 return HttpResponseRedirect('/detail')
#
#
#
#
#         return render(request,'detail.html',context=context)
#     else:
#         return HttpResponseRedirect('/message')
#

def detail(request):
    name = request.session.get('name')
    email = request.session.get('email')
    if asinDetail.objects.filter(email=email).filter(status=False).filter(extracted=True).exists():

        data1 = asinDetail.objects.filter(email=email).filter(status=False).filter(extracted=True)
        for i in data1:
            asin = i.asin

            if oldDetailAmazon.objects.filter(status=False).filter(asin=asin):
                #---->>>><<<<<-----#
                data2 = oldDetailAmazon.objects.get(asin=asin)
                #---->>>><<<<<-----#
                form1 = forms.form_newProductDetail()
                form2 = forms.form_oldProductDetail()
                form3 = forms.form_featureImage()

                data4 = ast.literal_eval(data2.old_desc)
                data7 = ast.literal_eval(data2.old_from_manufacture_h)
                data8 = ast.literal_eval(data2.old_from_manufacture_p)

                data9 = zip(data7,data8)

                #---->>>><<<<<-----#
                data5 = empDetail.objects.get(email=email)
                data6 = asinDetail.objects.filter(email=email)
                total_asin_allocated = data6.count()

                #---->>>><<<<<-----#

                print(type(data2.old_desc))

                context = {
                    'data2':data2,
                    'name':name,
                    'data4':data4,
                    'form1':form1,
                    'form2':form2,
                    'form3':form3,

                    'total_asin_allocated':total_asin_allocated,
                    'data5':data5,
                    'data7':data7,
                    'data8':data8,
                    'data9':data9

                    }
                if request.method == 'POST':
                    newtitle = request.POST.get('newtitle')
                    form1 = forms.form_newProductDetail(request.POST)
                    form2 = forms.form_oldProductDetail(request.POST)
                    form3 = forms.form_featureImage(request.POST)
                    sendme = asinDetail.objects.get(asin=asin)
                    sendme1 = oldDetailAmazon.objects.get(asin=asin)



                    if form1.is_valid() and form2.is_valid() and form3.is_valid():
                        obj = form1.save(commit=False)
                        obj.asin=asin
                        obj.save()

                        obj1 = form2.save(commit=False)
                        obj1.asin=asin
                        obj1.current_Title=data2.old_name
                        obj1.revised_Title=form1.cleaned_data['title']
                        obj1.save()

                        #--------->>>>>featureImage<<<<<<<<-----------

                        obj2 = form3.save(commit=False)

                        #---------ASIN----------------1
                        obj2.asin=asin
                        #=========ASIN-END=============


                        #----------connectivity_ports-------------2
                        hdmi = request.POST.get('hdmi1')
                        usb = request.POST.get('usb1')
                        vga = request.POST.get('vga1')

                        if(hdmi!="" and usb=="" and vga==""):
                            value3 = hdmi+" HDMI Ports"
                        elif(hdmi=="" and usb!="" and vga==""):
                            value3 = usb+" USB Ports"
                        elif(hdmi=="" and usb=="" and vga!=""):
                            value3 = vga+" VGA Ports"
                        elif(hdmi!="" and usb!="" and vga==""):
                            value3 = hdmi+" HDMI Ports |"+usb+" USB Ports"
                        elif(hdmi!="" and usb=="" and vga!=""):
                            value3 = hdmi+" HDMI Ports |"+vga+" VGA Ports"
                        elif(hdmi=="" and usb!="" and vga!=""):
                            value3 = usb+" USB Ports |"+vga+" VGA Ports"
                        elif(hdmi!="" and usb!="" and vga!=""):
                            value3 = hdmi+" HDMI Ports |"+usb+" USB Ports |"+vga+" VGA Ports"
                        else:
                            value3 = ""
                        obj2.connectivity_ports=value3
                        #==============connectivity_ports_END===============


                        #-----------------sound_output---------------------3
                        s1 = request.POST.get('s11')
                        s2 = request.POST.get('s22')
                        s3 = request.POST.get('s33')
                        s4 = request.POST.get('s44')

                        if(s1!="" and s2=="" and s3=="" and s4==""):
                            value5 = s1
                        elif(s1!="" and s2!="" and s3=="" and s4==""):
                            value5 = s1+" with "+s2
                        elif(s1!="" and s2!="" and s3!="" and s4==""):
                            value5 = s1+" with "+s2+", "+s3
                        elif(s1!="" and s2!="" and s3!="" and s4!=""):
                            value5 = s1+" with "+s2+", "+s3+", "+s4

                        else:
                            value5 = ""
                        obj2.sound_output=value5
                        #=================sound_output_END====================



                        #----------------display_and_refresh_rate---------------4
                        ds1 = request.POST.get('ds11')
                        ds2 = request.POST.get('ds22')
                        ds3 = request.POST.get('ds3')
                        ds4 = request.POST.get('ds4')
                        ds5 = request.POST.get('ds5')
                        rf = request.POST.get('refresh')

                        if(ds1!="" and ds2=="" and rf==""):
                            value4 = ds1
                        elif(ds1!="" and ds2=="" and rf!=""):
                            value4 = ds1+" | "+rf
                        elif(ds1!="" and ds2!="" and rf==""):
                            value4 = ds1+" | "+ds2
                        elif(ds1!="" and ds2!="" and rf!=""):
                            value4 = ds1+" | "+ds2+" | "+rf
                        elif(ds1=="" and ds2=="" and rf==""):
                            value4 = ""

                        obj2.display_and_refresh_rate=value4
                        #===============display_and_refresh_rate_END============


                        #--------------screen_size_resolution-------------------5
                        resolution = request.POST.get('resolution')
                        screen_size = request.POST.get('screen')
                        value2 = screen_size+" inches | "+resolution
                        obj2.screen_size_resolution=value2
                        #=============screen_size_resolution_END================



                        #-------------warranty----------------6
                        value7 = "Yes | Brand provided"
                        obj2.warranty=value7
                        #===========warranty-END==============



                        #-------------smart_tv-----------------7
                        smart = request.POST.get('smart')
                        fist = request.POST.get('fist')
                        if smart == "NO":
                            value6 = "No"
                        else:
                            value6 = "Yes | "+fist
                        obj2.smart_tv=value6
                        #==============smart_tv=================7


                        obj2.save()























                        #=============>>>>>featureImageEnd<<<<<<<<=============












                        data7 = empDetail.objects.get(email=email)
                        data7.asin_done_c = F('asin_done_c') + 1
                        data7.save()

                        if sendme.status == False:
                            sendme.status=True
                            sendme.save()

                        if sendme1.status == False:
                            sendme1.status=True
                            sendme1.save()
                        return HttpResponseRedirect('/detail')





                return render(request,'detail.html',context=context)
    else:
        return HttpResponseRedirect('/message')
